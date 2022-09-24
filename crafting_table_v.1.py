from urllib.request import urlopen
from bs4 import BeautifulSoup
from openpyxl import Workbook
from datetime import datetime
import schedule
import time
import sys
import os


class website:
    def __init__(self, name, link):
        self.name = name
        self.link = link


U1 = website("한국대학교", "URL")
U2 = website("안암대학교", "URL")

U_list = [U1,U2]

blank = [""]

def main():
    now = datetime.now()
    print("\n\n크롤링 시작, 시작시간 : " + now.strftime('%m-%d %H:%M:%S'))

    wb = Workbook()
    ws = wb.active
    ws.title = "sheet1"

    try:
        for i in range(len(U_list)):
            now = datetime.now()
            print("\n" + U_list[i].name + " 경쟁률 크롤링을 시작합니다.")

            sheet = wb.create_sheet(U_list[i].name + "경쟁률")

            html = urlopen(U_list[i].link)
            bsObject = BeautifulSoup(html, "html.parser")

            f = open(U_list[i].name + " " + now.strftime('%m-%d %H:%M:%S') + ".html", 'w')
            f.write(str(bsObject))
            f.close()

            list = []
            for link in bsObject.find_all("td"):
                list.append((link.text.strip()))

            p = 1
            q = 1
            plag = 0
            for j in list:
                for k in blank:
                    if j in k:
                        if plag == 0:
                            plag += 1
                        elif plag == 1:
                            plag = 0
                if plag == 0:
                    sheet.cell(column=q, row=p, value=j)
                    if q == 4:
                        q = 1
                        p += 1
                    else:
                        q += 1
                    plag = 0

            now = datetime.now()
            print(U_list[i].name + " 크롤링 성공! 성공시각 : " + now.strftime('%m-%d %H:%M:%S'))

        wb.save('result_' + now.strftime('%m-%d %H:%M:%S') + '.xlsx')


    except:
        now = datetime.now()
        print("\n크롤링 실패!, 프로그램 재시작 필요.")
        print("이 오류로그는 " + now.strftime('%m-%d %H:%M:%S') + "에 표시되었습니다.")
        sys.exit()

    os.execv(sys.executable, ['python'] + sys.argv)


if __name__ == "__main__":
    schedule.every().day.at("00:00").do(main)

    while True:
        schedule.run_pending()
        time.sleep(1)